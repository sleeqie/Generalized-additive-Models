from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import json

# starting function
# NOTE: both function parameters must be strings when the function is invoked
def readFiles(path, worksheet):
    # loading workbook and worksheet
    wb = load_workbook(path)
    ws = wb[worksheet]

    # getting rows and columns
    rows = int(input("enter number of rows: "))
    cols = int(input("enter number of columns: "))

    # initialization for json format
    json_list = []
    

    # getting data from specified rows and columns
    for row in range(1, rows+1):
        for col in range(1, cols+1):
            char = get_column_letter(col)
            data_value = ws[char + str(row)].value

            # excepting null values
            if ws[char + str(row)].value == None:
                continue

            # printing data onto the terminal and adding to json_list
            else:
                json_list.append(data_value)
                print(ws[char + str(row)].value)
    # saving file. This is optional
    wb.save('newfile.xlsx')


    # writing data to a json file called data.json
    j = json.dumps(json_list)
    with open('data.json', 'w') as f:
        f.write(j)


# invoking function
readFiles("pp.xlsx", "Blanching")